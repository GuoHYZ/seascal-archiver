# -*- coding: utf-8 -*-
"""
XLSX 电子表格转 TXT 模块。

功能：将 Excel 文件（.xlsx）转换为纯文本文件。
- 每个工作表（Sheet）分别输出
- 表格转换为 Markdown 格式保留结构
- 空单元格记为空格，保持列对齐
- 图片忽略不处理

统一接口：convert(source_path, dest_path) -> int
"""

import sys
from pathlib import Path
import traceback

try:
    from rag_utils import escape_md_cell
    import rag_config as _cfg
except ImportError:
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from rag_utils import escape_md_cell
    import rag_config as _cfg

try:
    import openpyxl
except ImportError:
    openpyxl = None


def _sheet_to_markdown(ws, max_rows=None, max_cols=100):
    """
    将 openpyxl 工作表转换为 Markdown 表格字符串。

    ws:      工作表对象
    max_rows: 最大读取行数。
              None  → 使用 rag_config.XLSX_MAX_ROWS（0 = 不限制）
              0     → 不限制，读取全部行
              N > 0 → 限制读取前 N 行（超限会打印警告）
    max_cols: 最大读取列数
    返回: Markdown 表格字符串，或空字符串
    """
    if max_rows is None:
        max_rows = _cfg.XLSX_MAX_ROWS

    actual_max_row = (
        min(ws.max_row or 0, max_rows) if max_rows > 0 else (ws.max_row or 0)
    )
    actual_max_col = min(ws.max_column or 0, max_cols)

    # 检测是否超限
    if max_rows > 0 and ws.max_row and ws.max_row > max_rows:
        print(
            "  [警告] Sheet '{}' 共 {} 行，仅读取前 {} 行".format(
                ws.title, ws.max_row, max_rows
            )
        )

    rows_data = []

    for row in ws.iter_rows(
        min_row=1,
        max_row=actual_max_row,
        max_col=actual_max_col,
        values_only=True,
    ):
        row_data = [escape_md_cell(cell) for cell in row]
        # 跳过全空行
        if any(cell for cell in row_data):
            rows_data.append(row_data)

    if not rows_data:
        return ""

    # 确定列数
    col_count = max(len(r) for r in rows_data)
    if col_count == 0:
        return ""

    # 补齐列（不一致的行用空字符串填充）
    for row_data in rows_data:
        while len(row_data) < col_count:
            row_data.append("")

    lines = []

    # 表头行
    header = rows_data[0]
    lines.append("| " + " | ".join(header) + " |")

    # 分隔行
    lines.append("|" + "|".join([" --- " for _ in range(col_count)]) + "|")

    # 数据行
    for row_data in rows_data[1:]:
        lines.append("| " + " | ".join(row_data) + " |")

    return "\n".join(lines)


def xlsx_to_txt(source_path, dest_path):
    """
    将 XLSX 文件转换为 TXT 文件。

    source_path: str 或 Path — 源文件路径
    dest_path:   str 或 Path — 目标 TXT 文件路径
    """
    if openpyxl is None:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    source_path = Path(source_path)
    dest_path = Path(dest_path)

    wb = openpyxl.load_workbook(str(source_path), data_only=True, read_only=True)

    dest_path.parent.mkdir(parents=True, exist_ok=True)

    with open(dest_path, "w", encoding="utf-8", errors="replace") as f:
        f.write(f"文件：{source_path.name}\n")
        f.write("=" * 60 + "\n")

        sheet_count = 0
        for ws in wb.worksheets:
            sheet_count += 1
            f.write(f"\n---------- Sheet: {ws.title} ----------\n\n")

            try:
                md_table = _sheet_to_markdown(ws)
                if md_table:
                    f.write(md_table + "\n")
                else:
                    f.write("(Sheet 为空)\n")
            except Exception as e:
                f.write(f"[错误] Sheet 处理失败: {e}\n")

        f.write(f"\n{'=' * 60}\n")
        f.write(f"共 {sheet_count} 个工作表\n")

    wb.close()


def convert(source_path, dest_path):
    """
    统一转换接口。

    source_path: str 或 Path — 源 XLSX 文件路径
    dest_path:   str 或 Path — 目标 TXT 文件路径
    返回: int — 提取的总字符数
    """
    xlsx_to_txt(source_path, dest_path)

    dest_path = Path(dest_path)
    if dest_path.exists():
        return len(dest_path.read_text(encoding="utf-8", errors="replace"))
    return 0


# ============================================================
# 独立运行：批量转换当前目录下指定文件夹的 XLSX 文件
# ============================================================
if __name__ == "__main__":
    input_dir = Path("./XLSX")
    output_dir = Path("./TXT")

    if not input_dir.is_dir():
        print(f"[错误] 输入目录不存在: {input_dir}")
    else:
        output_dir.mkdir(exist_ok=True)
        xlsx_files = list(input_dir.glob("*.xlsx"))
        xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]

        if not xlsx_files:
            print(f"[警告] 目录 {input_dir} 中未找到任何 XLSX 文件")
        else:
            success, failed = 0, 0
            total = len(xlsx_files)
            for i, xlsx in enumerate(xlsx_files, start=1):
                txt = output_dir / (xlsx.stem + ".txt")
                try:
                    char_count = convert(xlsx, txt)
                    success += 1
                    print(f"[{i}/{total}] [OK] {xlsx.name} ({char_count} 字符)")
                except Exception as e:
                    failed += 1
                    print(f"[{i}/{total}] [FAIL] {xlsx.name}")
                    print(f"       原因: {e}")
                    traceback.print_exc()
            print(f"\n处理结束 - 共 {total} 个文件")
            print(f"成功: {success}, 失败: {failed}")